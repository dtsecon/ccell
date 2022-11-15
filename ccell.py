#
#  ccell - change a cell range in a workbook
#
#  View and change a cell range in a worksheet 
#  For more information of ccell tool, bug reports, future suggestions 
#  and track changes visit: https://github.com/dtsecon/ccell
#
#  Copyright (C) 2022, Dimitris Economou (dimitris.s.economou@gmail.com)
# 
#  This file is part of ccell.
# 
#  ccell is free software: you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation, either version 3 of the License, or
#  (at your option) any later version.
# 
#  modio is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  You should have received a copy of the GNU General Public License
#  along with modio.  If not, see <https://www.gnu.org/licenses/>.
  

import getopt
import sys
import os

try:
    from openpyxl import Workbook, load_workbook
except ModuleNotFoundError as err:
    print("Error: install module openpyxl (%s)" % (err))
    sys.exit(os.EX_NOTFOUND)

# print a sheet range in table format
def printRange(range):

    # find max length for each column of range (initial length 4)
    maxlen = [4] * len(range[0])
    for r in range:
        for i, v in enumerate(r):
            if len(str(v.value)) > maxlen[i]:
                maxlen[i] = len(str(v.value))

    # print range in table (row-columnt) format
    print()
    if type(range) in [list, tuple]:
        for i, r in enumerate(range):
            print(f"%3d|" % (i), end = " ")
            for j, v in enumerate(r):
                if v.data_type == 's' or v.value == None:
                    fmt = "%" + str(maxlen[j]) + "s"
                else:
                    fmt = "%" + str(maxlen[j]) + ".1f"
                print(f"{fmt}|" % (v.value), end = " ")
            print()

def usage():
    print();
    print("Usage: %s [OPTIONS]..." % (sys.argv[0]))
    print("--(h)elp                    print usage")
    print("--(f)ile         <filename> the filename of the workbook")
    print("--shee(t)      <sheet name> the sheet name in the workbook")
    print("--(i)ndex     <sheet index> the sheet index in the workbook")
    print("--(c)ell          <address> the cell or range of cells using column-row notation e.g. A12 or A2:E12")
    print("                            if cell not provided, entire sheet is active range")
    print("--(w)rite           <value> write a new value in cell or cell range")
    print("--stri(p)                   strip text value from leading and trailing spaces")
    print("--(s)ave         <filename> save workbook as a different file")
    print("--d(r)y                     run without saving to file")
    print("--(l)ist_sheets             list the available sheets in workbook")

def main():
    try:
        opts, args = getopt.getopt(sys.argv[1:], 
                                   "hf:t:i:c:w:ps:rlvd", 
                                   ["help", "file=", "sheet=", "index=", "cell=", "write=", "strip", "save=", "dry", "list_sheets", "verbose", "debug"]
                    )
    except getopt.GetoptError as err:

        # print help information and exit:
        print(err)
        usage()
        sys.exit(os.EX_USAGE)

    # if no options passed, print usage and exit
    if not opts:
        usage()
        sys.exit(os.EX_USAGE) 

    # initialize states
    file = None
    sheet = None
    index = None
    cell = None
    write = None
    strip = False
    save = None
    dry = None
    listSheets = False
    verbose = False
    debug = False

    # go through all options
    for o, a in opts:
        if o in ("-d", "--debug"):
            debug = True
        elif o in ("-v", "--verbose"):
            verbose = True
        elif o in ("-f", "--file"):
            file = a
        elif o in ("-t", "--sheet"):
            sheet = a
        elif o in ("-i", "--index"):
            try:
                index = int(a)
            except ValueError as err:
                print("Error: option -i <index> (%s)" % (err))
                sys.exit(os.EX_USAGE)
        elif o in ("-c", "--cell"):
            cell = a
        elif o in ("-w", "--write"):
            write = a
        elif o in ("-p", "--strip"):
            strip = True
        elif o in ("-s", "--save"):
            save = a
        elif o in ("-r", "--dry"):
            dry = True
        elif o in ("-l", "--list_sheets"):
            listSheets = True
        else:
            assert False, "unhandled option"
        
    
    if debug:
        for o, a in opts:
            print(o, a)
        print()
        print("state variables:")
        print("file: %s" % (file))
        print("sheet: %s" % (sheet))
        print("index: %d" % (index))
        print("cell: %s" % (cell))
        print("write: %s" % (write))


    # for any arguments without an option print a message and exit
    if args:
        for a in args:
            print("Unrecognized argument", a)
        usage()
        sys.exit(os.EX_USAGE)
        

    # check if a filename has been defined
    if file:
        wb = load_workbook(file)
    else:
        print("Filename is missing (use -f <filename>)")
        usage()
        sys.exit(os.EX_USAGE)

    # check if list option has been given
    if listSheets:    
        for s in wb.worksheets:
            print(wb.index(s), s.title)
        sys.exit(os.EX_USAGE)

    # check if sheet name or index has been defined
    if sheet:
        try:
            ws = wb[sheet]
        except KeyError as err:
            print("Error: option -t <sheet name> (%s)" % (err))
            sys.exit(os.EX_USAGE)
    elif index != None:
        try:
            ws = wb.worksheets[index]
        except IndexError as err:
            print("Error: option -i <sheet index> (%s)" % (err))
            sys.exit(os.EX_USAGE)
    else:
        print("Sheet name or index is missing (use -s <sheet name> or -i <sheet index>)")
        usage()
        sys.exit(os.EX_USAGE)
        
    # check if a cell address has been defined and get the range
    if cell:
        try:
            c = ws[cell]
        except ValueError as err:
            print("Error: option -c <address> (%s)" % (err))
            sys.exit(os.EX_USAGE)
    # otherwise get the whole sheet
    else:
        c = ws["A1:" + 
               ws.cell(row=ws.max_row, column=ws.max_column).column_letter + 
               str(ws.cell(row=ws.max_row, column=ws.max_column).row)
            ]
        #print("Cell address is missing (use -c <address>)")
        #usage()
        #sys.exit(os.EX_USAGE)
        
    # print value of cell
    if type(c) in [list, tuple]:
        printRange(c)
    else:
        if c.data_type == 's':
            print("cell %s!%s value: %s type: %s length: %s" % (ws.title, cell, c.value, c.data_type, len(c.value)))
        else:
            print("cell %s!%s value: %s type: %s" % (ws.title, cell, c.value, c.data_type))

    # check if write value has been defined
    # check cell data type and handle value accordingly
    if write:
        if type(c) in [list, tuple]:
            for r in c:
                for v in r:
                    if v.data_type == 's':
                        v.value = write
                    elif v.data_type == 'n':
                        try:
                            v.value = float(write)
                        except ValueError as err:
                            print("Error: option -w <value> (%s)" % (err))
                            sys.exit(os.EX_USAGE)
            printRange(c)
        elif c.data_type == 's':
            c.value = write
            print("cell %s!%s new value: %s type: %s" % (ws.title, cell, c.value, c.data_type))
        elif c.data_type == 'n':
            try:
                c.value = float(write)
            except ValueError as err:
                print("Error: option -w <value> (%s)" % (err))
                sys.exit(os.EX_USAGE)
            print("cell %s!%s new value: %s type: %s" % (ws.title, cell, c.value, c.data_type))

    # Strip a cell with text value from leading and trailing spaces
    if strip:
        if type(c) in [list, tuple]:
            for r in c:
                for v in r:
                    if v.data_type == 's':
                        v.value = v.value.strip() 
            printRange(c)            
        elif c.data_type == 's':
            c.value = c.value.strip()
            print("cell %s!%s value: %s type: %s length: %s" % (ws.title, cell, c.value, c.data_type, len(c.value)))
        

    # if a dry run exit with success
    if dry:
        sys.exit(os.EX_OK)

    # save to a different file if write has been defined and exit with success
    if save and (write or strip):
        print("Saving to file %s" % (save))
        wb.save(save)
        sys.exit(os.EX_OK)

    # save to file if write has been defined
    if write or strip:
        print("Saving to file %s" % (file))
        wb.save(file)

    # exit with success
    sys.exit(os.EX_OK)

if __name__ == "__main__":
    main()